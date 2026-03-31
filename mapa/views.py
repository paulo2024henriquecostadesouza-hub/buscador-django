import json
import unicodedata
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone

from .models import Servico, ConsultaProgramada, ImportacaoRelatorio, ProgramacaoServico
from .importador import importar_relatorio_flip, importar_consultas_programadas


# Limites da área de atuação (São Paulo metropolitana)
# Pontos fora dessa caixa são ocultados do mapa
LAT_MIN, LAT_MAX = -24.5, -22.5
LNG_MIN, LNG_MAX = -47.5, -45.0

CORES_FONTE = {
    'SAC':      '#4fc3f7',   # azul claro
    'OUVIDORIA':'#81c784',   # verde
    'ACIC':     '#ff8a65',   # laranja
    'BFS':      '#ffd54f',   # amarelo
    'CNC':      '#ce93d8',   # roxo
}
COR_PADRAO = '#e94560'


def _normalizar_busca(texto):
    """Remove acentos e converte para lowercase."""
    nfkd = unicodedata.normalize('NFKD', str(texto))
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower()


def index(request):
    """Página principal com o mapa de calor."""
    tipos  = list(Servico.objects.values_list('tipo_servico', flat=True).distinct().order_by('tipo_servico'))
    fontes = list(Servico.objects.values_list('fonte', flat=True).distinct().order_by('fonte'))

    # Status globais (para exibição inicial sem filtro de fonte)
    status_list = list(Servico.objects.values_list('status', flat=True).distinct().order_by('status'))

    # Status agrupados por fonte — usados no filtro dinâmico do JS
    status_por_fonte = {}
    for fonte in fontes:
        sts = list(
            Servico.objects.filter(fonte=fonte)
            .values_list('status', flat=True)
            .distinct().order_by('status')
        )
        status_por_fonte[fonte] = sts

    context = {
        'tipos': tipos,
        'status_list': status_list,
        'fontes': fontes,
        'status_por_fonte_json': json.dumps(status_por_fonte, ensure_ascii=False),
        'cores_fonte': json.dumps(CORES_FONTE),
        'cor_padrao': COR_PADRAO,
        'total_servicos': Servico.objects.filter(latitude__isnull=False).count(),
        'total_programados': ConsultaProgramada.objects.filter(latitude__isnull=False).count(),
    }
    return render(request, 'mapa/index.html', context)


def _feito_q():
    """Retorna Q para serviços considerados 'feitos'."""
    return (
        Q(data_finalizacao__isnull=False) |
        Q(data_execucao__isnull=False) |
        Q(status__icontains='finaliz') |
        Q(status__icontains='execut') |
        Q(status__icontains='conclu') |
        Q(status__icontains='realiz')
    )


def dados_mapa(request):
    """API JSON com os pontos para o mapa."""
    tipo_filtro   = request.GET.get('tipo', '')
    status_filtro = request.GET.get('status', '')
    fonte_filtro  = request.GET.get('fonte', '')
    busca         = request.GET.get('busca', '').strip()
    mes           = request.GET.get('mes', '').strip()   # formato YYYY-MM
    resolucao     = request.GET.get('resolucao', '').strip()

    servicos_qs = Servico.objects.filter(
        latitude__isnull=False, longitude__isnull=False,
        latitude__gte=LAT_MIN,  latitude__lte=LAT_MAX,
        longitude__gte=LNG_MIN, longitude__lte=LNG_MAX,
    )

    if mes:
        try:
            ano, m = mes.split('-')
            servicos_qs = servicos_qs.filter(
                Q(data_execucao__year=int(ano), data_execucao__month=int(m)) |
                Q(data_finalizacao__year=int(ano), data_finalizacao__month=int(m))
            )
        except (ValueError, AttributeError):
            pass

    if tipo_filtro:
        servicos_qs = servicos_qs.filter(tipo_servico__icontains=tipo_filtro)
    if status_filtro:
        servicos_qs = servicos_qs.filter(status__icontains=status_filtro)
    if fonte_filtro:
        servicos_qs = servicos_qs.filter(fonte__iexact=fonte_filtro)
    if busca:
        busca_norm = _normalizar_busca(busca)
        servicos_qs = servicos_qs.filter(
            Q(endereco_normalizado__icontains=busca_norm) |
            Q(numero_os__icontains=busca)
        )

    if resolucao == 'feito':
        servicos_qs = servicos_qs.filter(_feito_q())
    elif resolucao == 'pendente':
        servicos_qs = servicos_qs.exclude(_feito_q())

    # Agrupa por coordenada aproximada (~100m)
    pontos = {}
    for s in servicos_qs.values('latitude', 'longitude', 'tipo_servico', 'status',
                                 'endereco', 'data_execucao', 'data_finalizacao',
                                 'numero_os', 'fonte', 'fiscal'):
        chave = (round(s['latitude'], 4), round(s['longitude'], 4))
        if chave not in pontos:
            pontos[chave] = {
                'lat': chave[0],
                'lng': chave[1],
                'endereco': s['endereco'],
                'servicos': [],
                'quantidade': 0,
                'fonte': s['fonte'],
            }
        pontos[chave]['servicos'].append({
            'os':               s['numero_os'],
            'tipo':             s['tipo_servico'],
            'status':           s['status'],
            'fonte':            s['fonte'],
            'data':             s['data_execucao'].strftime('%d/%m/%Y') if s['data_execucao'] else '',
            'data_finalizacao': s['data_finalizacao'].strftime('%d/%m/%Y') if s['data_finalizacao'] else '',
            'fiscal':           s['fiscal'] or '',
        })
        pontos[chave]['quantidade'] += 1
        # Se houver múltiplas fontes no mesmo ponto, usa a mais frequente (última)
        if pontos[chave]['fonte'] != s['fonte']:
            pontos[chave]['fonte'] = s['fonte']

    # Consultas programadas
    programados = []
    for c in ConsultaProgramada.objects.filter(latitude__isnull=False, longitude__isnull=False):
        programados.append({
            'lat': c.latitude,
            'lng': c.longitude,
            'endereco': c.endereco,
            'tipo': c.tipo_servico,
            'data_programada': c.data_programada.strftime('%d/%m/%Y'),
            'observacao': c.observacao,
        })

    return JsonResponse({
        'pontos': list(pontos.values()),
        'programados': programados,
    })


def exportar_excel(request):
    """Exporta os serviços filtrados para Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tipo_filtro = request.GET.get('tipo', '')
    status_filtro = request.GET.get('status', '')
    fonte_filtro = request.GET.get('fonte', '')
    busca = request.GET.get('busca', '').strip()

    qs = Servico.objects.all().order_by('-data_execucao')
    if tipo_filtro:
        qs = qs.filter(tipo_servico__icontains=tipo_filtro)
    if status_filtro:
        qs = qs.filter(status__icontains=status_filtro)
    if fonte_filtro:
        qs = qs.filter(fonte__iexact=fonte_filtro)
    if busca:
        busca_norm = _normalizar_busca(busca)
        qs = qs.filter(
            Q(endereco_normalizado__icontains=busca_norm) |
            Q(numero_os__icontains=busca)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Serviços'

    cabecalho = ['Número OS', 'Tipo de Serviço', 'Status', 'Fonte',
                 'Endereço', 'Latitude', 'Longitude',
                 'Data Execução', 'Data Abertura']

    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(bold=True, color='e94560')

    for col, titulo in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for s in qs.values('numero_os', 'tipo_servico', 'status', 'fonte',
                        'endereco', 'latitude', 'longitude',
                        'data_execucao', 'data_abertura'):
        ws.append([
            s['numero_os'],
            s['tipo_servico'],
            s['status'],
            s['fonte'],
            s['endereco'],
            s['latitude'],
            s['longitude'],
            s['data_execucao'].strftime('%d/%m/%Y') if s['data_execucao'] else '',
            s['data_abertura'].strftime('%d/%m/%Y') if s['data_abertura'] else '',
        ])

    larguras = [14, 55, 25, 12, 60, 14, 14, 16, 16]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = larg

    data_hoje = timezone.localdate().strftime('%Y%m%d')
    nome_arquivo = f'servicos_{data_hoje}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)
    return response


def programacao_mapa(request):
    """
    API: retorna pontos da programação agrupados por setor.
    Cada ponto traz as próximas datas agendadas.
    Filtra por subprefeitura e/ou tipo de serviço se informado.
    """
    from datetime import date
    import re

    sub_filtro  = request.GET.get('sub', '')
    tipo_filtro = request.GET.get('tipo_prog', '')
    busca       = request.GET.get('busca', '').strip()
    data_filtro = request.GET.get('data', '').strip()   # formato YYYY-MM-DD

    qs = ProgramacaoServico.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    )
    if sub_filtro:
        qs = qs.filter(subprefeitura__iexact=sub_filtro)
    if tipo_filtro:
        qs = qs.filter(tipo_nome__icontains=tipo_filtro)

    # Filtro por data específica: mantém só setores que têm essa data em opcoes_data
    data_str_busca = ''
    if data_filtro:
        try:
            import datetime as _dt
            dt = _dt.datetime.strptime(data_filtro, '%Y-%m-%d')
            data_str_busca = dt.strftime('%d/%m/%Y')
            qs = qs.filter(opcoes_data__icontains=data_str_busca)
        except ValueError:
            pass

    busca_norm = _normalizar_busca(busca) if busca else ''
    if busca:
        qs = qs.filter(
            Q(logradouro__icontains=busca) |
            Q(setor__icontains=busca) |
            Q(nome_local__icontains=busca)
        )

    hoje = date.today()

    # Agrupa por setor
    # Quando há busca por logradouro: posiciona o marcador no centro dos pontos
    # que batem com a busca (não no primeiro ponto geral do setor)
    setores = {}
    # Acumula pontos que casam com a busca para calcular centroide por setor
    match_coords = {}  # setor -> lista de [lat, lng] que bateram com a busca

    for p in qs.values('latitude', 'longitude', 'setor', 'tipo_nome',
                        'subprefeitura', 'logradouro', 'nome_local',
                        'referencia', 'horario_inicio', 'horario_fim',
                        'opcoes_data', 'data_inicio', 'data_fim',
                        'qtd_operadores', 'qtd_veiculos'):

        setor = p['setor']

        # Acumula coordenadas que casam com a busca
        if busca_norm:
            log_norm = _normalizar_busca(p['logradouro'] or '')
            loc_norm = _normalizar_busca(p['nome_local'] or '')
            if busca_norm in log_norm or busca_norm in loc_norm or busca_norm in (p['setor'] or '').lower():
                match_coords.setdefault(setor, []).append([p['latitude'], p['longitude']])

        if setor not in setores:
            # Calcula próximas datas
            proximas = []
            opcoes = p['opcoes_data'] or ''
            opcoes = re.sub(r'\s*[-–]\s*', ';', opcoes)
            for ds in opcoes.split(';'):
                ds = ds.strip()
                for fmt in ('%d/%m/%Y', '%d/%m/%y'):
                    try:
                        d = __import__('datetime').datetime.strptime(ds, fmt).date()
                        if d >= hoje:
                            proximas.append(d)
                        break
                    except ValueError:
                        continue
            proximas.sort()

            setores[setor] = {
                'lat':            p['latitude'],
                'lng':            p['longitude'],
                'setor':          setor,
                'tipo':           p['tipo_nome'],
                'subprefeitura':  p['subprefeitura'],
                'logradouro':     p['logradouro'],
                'nome_local':     p['nome_local'] or '',
                'horario':        f"{p['horario_inicio']} – {p['horario_fim']}",
                'qtd_operadores': p['qtd_operadores'],
                'qtd_veiculos':   p['qtd_veiculos'],
                'data_inicio':    p['data_inicio'].strftime('%d/%m/%Y') if p['data_inicio'] else '',
                'data_fim':       p['data_fim'].strftime('%d/%m/%Y')    if p['data_fim']    else '',
                'proximas_datas': [d.strftime('%d/%m/%Y') for d in proximas[:6]],
                'total_pontos':   1,
                'busca_logradouro': busca if busca else '',
            }
        else:
            setores[setor]['total_pontos'] += 1

    # Reposiciona marcador no centroide dos pontos que casaram com a busca
    if busca_norm:
        for setor, coords in match_coords.items():
            if setor in setores and coords:
                lat_c = sum(c[0] for c in coords) / len(coords)
                lng_c = sum(c[1] for c in coords) / len(coords)
                setores[setor]['lat'] = lat_c
                setores[setor]['lng'] = lng_c

    return JsonResponse({'setores': list(setores.values())})


def segmentos_do_dia(request, data_str):
    """
    Retorna todos os segmentos (polylines) de todos os setores
    que têm a data informada em opcoes_data.
    data_str formato: YYYY-MM-DD
    """
    import datetime as _dt

    try:
        dt = _dt.datetime.strptime(data_str, '%Y-%m-%d')
        data_br = dt.strftime('%d/%m/%Y')
    except ValueError:
        return JsonResponse({'setores': []})

    qs = (
        ProgramacaoServico.objects
        .filter(latitude__isnull=False, opcoes_data__icontains=data_br)
        .order_by('setor', 'id')
        .values('setor', 'tipo_nome', 'subprefeitura', 'logradouro', 'latitude', 'longitude')
    )

    # Agrupa por setor → segmentos por logradouro
    from collections import defaultdict
    setores_dict = {}     # setor → {tipo, sub, segmentos_dict}
    seg_dict = {}         # (setor, logradouro) → lista de coords

    for p in qs:
        setor    = p['setor']
        log      = p['logradouro'] or ''
        coord    = [p['latitude'], p['longitude']]

        if setor not in setores_dict:
            setores_dict[setor] = {
                'setor':         setor,
                'tipo':          p['tipo_nome'],
                'subprefeitura': p['subprefeitura'],
                'segmentos':     [],
            }

        chave = (setor, log)
        if chave not in seg_dict:
            seg_dict[chave] = []
        seg_dict[chave].append(coord)

    # Monta segmentos com mínimo de 2 pontos
    for (setor, log), coords in seg_dict.items():
        if len(coords) >= 2:
            setores_dict[setor]['segmentos'].append({
                'logradouro': log,
                'coords':     coords,
            })

    # Filtra setores que têm pelo menos 1 segmento
    resultado = [s for s in setores_dict.values() if s['segmentos']]

    return JsonResponse({'data': data_br, 'setores': resultado})


def setor_pontos(request, setor):
    """
    Retorna os pontos de um setor agrupados por logradouro.
    Cada segmento = uma rua = uma polyline separada no mapa.
    """
    pontos = (
        ProgramacaoServico.objects
        .filter(setor=setor, latitude__isnull=False)
        .order_by('id')
        .values('latitude', 'longitude', 'logradouro')
    )

    # Agrupa pontos consecutivos por logradouro
    segmentos = []
    seg_atual = None
    log_atual = None

    for p in pontos:
        log = p['logradouro'] or ''
        coord = [p['latitude'], p['longitude']]

        if log != log_atual:
            # Começa novo segmento
            if seg_atual and len(seg_atual['coords']) >= 2:
                segmentos.append(seg_atual)
            seg_atual = {'logradouro': log, 'coords': [coord]}
            log_atual = log
        else:
            seg_atual['coords'].append(coord)

    # Adiciona o último segmento
    if seg_atual and len(seg_atual['coords']) >= 2:
        segmentos.append(seg_atual)

    return JsonResponse({
        'setor':     setor,
        'segmentos': segmentos,
    })


def regioes_geojson(request):
    """Serve o GeoJSON das regionais de atuação."""
    import json
    from pathlib import Path
    from django.conf import settings

    caminho = Path(settings.BASE_DIR) / 'data' / 'regioes.json'
    if not caminho.exists():
        return JsonResponse({'type': 'FeatureCollection', 'features': []})
    data = json.loads(caminho.read_text(encoding='utf-8'))
    return JsonResponse(data)


def stats_mapa(request):
    """
    API: retorna estatísticas dos serviços para o painel lateral.
    Respeita os mesmos filtros de dados_mapa (mes, tipo, status, fonte, busca, resolucao).
    """
    from django.db.models import Count

    tipo_filtro   = request.GET.get('tipo', '')
    status_filtro = request.GET.get('status', '')
    fonte_filtro  = request.GET.get('fonte', '')
    busca         = request.GET.get('busca', '').strip()
    mes           = request.GET.get('mes', '').strip()
    resolucao     = request.GET.get('resolucao', '').strip()

    qs = Servico.objects.filter(
        latitude__isnull=False,
        latitude__gte=LAT_MIN,  latitude__lte=LAT_MAX,
        longitude__gte=LNG_MIN, longitude__lte=LNG_MAX,
    )

    if mes:
        try:
            ano, m = mes.split('-')
            qs = qs.filter(
                Q(data_execucao__year=int(ano), data_execucao__month=int(m)) |
                Q(data_finalizacao__year=int(ano), data_finalizacao__month=int(m))
            )
        except (ValueError, AttributeError):
            pass
    if tipo_filtro:
        qs = qs.filter(tipo_servico__icontains=tipo_filtro)
    if status_filtro:
        qs = qs.filter(status__icontains=status_filtro)
    if fonte_filtro:
        qs = qs.filter(fonte__iexact=fonte_filtro)
    if busca:
        busca_norm = _normalizar_busca(busca)
        qs = qs.filter(
            Q(endereco_normalizado__icontains=busca_norm) |
            Q(numero_os__icontains=busca)
        )

    if resolucao == 'feito':
        qs = qs.filter(_feito_q())
    elif resolucao == 'pendente':
        qs = qs.exclude(_feito_q())

    total = qs.count()

    # Feitos x Pendentes (usando o mesmo critério do _feito_q)
    feitos   = qs.filter(_feito_q()).count()
    pendentes = total - feitos

    # Por fonte (com breakdown feito/pendente)
    por_fonte_raw = list(
        qs.values('fonte').annotate(total=Count('id')).order_by('-total')
    )

    # Por tipo (top 10)
    por_tipo = list(
        qs.values('tipo_servico').annotate(total=Count('id')).order_by('-total')[:10]
    )

    # Por regional — agrupado por subprefeitura da programação
    por_regional = list(
        ProgramacaoServico.objects
        .values('subprefeitura')
        .annotate(setores=Count('setor', distinct=True), pontos=Count('id'))
        .order_by('subprefeitura')
    )

    return JsonResponse({
        'total':        total,
        'feitos':       feitos,
        'pendentes':    pendentes,
        'por_fonte':    por_fonte_raw,
        'por_tipo':     por_tipo,
        'por_regional': por_regional,
    })


def atualizar_flip(request):
    """
    Dispara a automação Selenium em segundo plano para baixar e importar
    os relatórios do Selimp diretamente no banco de dados.
    """
    if request.method != 'POST':
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(['POST'])

    import subprocess
    import sys
    from django.conf import settings

    data_inicio = request.POST.get('data_inicio', '01/04/2025').strip() or '01/04/2025'
    sem_limpar  = request.POST.get('sem_limpar') == '1'

    cmd = [sys.executable, 'manage.py', 'atualizar_flip', f'--data-inicio={data_inicio}']
    if sem_limpar:
        cmd.append('--sem-limpar')

    try:
        subprocess.Popen(
            cmd,
            cwd=str(settings.BASE_DIR),
            creationflags=subprocess.CREATE_NEW_CONSOLE,   # abre janela separada no Windows
        )
        messages.success(
            request,
            f'Automação iniciada! O Chrome será aberto e os dados importados automaticamente. '
            f'Período: {data_inicio} → hoje.'
        )
    except Exception as e:
        messages.error(request, f'Erro ao iniciar automação: {e}')

    return redirect('mapa:importar')


def reimportar_pasta(request):
    """Limpa todos os serviços e reimporta os arquivos da pasta Flip."""
    if request.method != 'POST':
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(['POST'])

    import os
    from django.conf import settings

    pasta = os.path.join(settings.BASE_DIR, 'Flip')
    if not os.path.isdir(pasta):
        messages.error(request, f'Pasta Flip não encontrada em: {pasta}')
        return redirect('mapa:importar')

    arquivos = [f for f in os.listdir(pasta)
                if f.lower().endswith('.csv') or f.lower().endswith('.xlsx')]

    if not arquivos:
        messages.error(request, 'Nenhum arquivo encontrado na pasta Flip.')
        return redirect('mapa:importar')

    # Limpa dados anteriores
    Servico.objects.all().delete()
    ImportacaoRelatorio.objects.filter(tipo='relatorio').delete()

    total_geral = 0
    erros = []
    for fname in sorted(arquivos):
        path = os.path.join(pasta, fname)
        try:
            with open(path, 'rb') as f:
                class ArquivoWrapper:
                    name = fname
                    def read(self_): return f.read()
                total = importar_relatorio_flip(ArquivoWrapper())
                total_geral += total
        except Exception as e:
            erros.append(f'{fname}: {e}')

    if erros:
        for err in erros:
            messages.error(request, f'Erro em {err}')

    messages.success(request, f'{total_geral} serviços reimportados de {len(arquivos)} arquivo(s) com encoding correto.')
    return redirect('mapa:importar')


def importar(request):
    """Tela para importar relatórios."""
    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo')
        tipo = request.POST.get('tipo', 'relatorio')

        if not arquivo:
            messages.error(request, 'Nenhum arquivo selecionado.')
            return redirect('mapa:importar')

        try:
            if tipo == 'relatorio':
                total = importar_relatorio_flip(arquivo)
                messages.success(request, f'{total} serviços importados com sucesso.')
            else:
                total = importar_consultas_programadas(arquivo)
                messages.success(request, f'{total} consultas programadas importadas.')
        except Exception as e:
            messages.error(request, f'Erro ao importar: {e}')

        return redirect('mapa:importar')

    historico = ImportacaoRelatorio.objects.all()[:10]
    return render(request, 'mapa/importar.html', {'historico': historico})
